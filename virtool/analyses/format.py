"""Functions and data to use for formatting Pathoscope and NuVs analysis results.

Formatted results are destined for API responses or CSV/Excel formatted file
downloads.
"""

import csv
import io
import statistics
from collections import defaultdict
from typing import Any

import openpyxl.styles
import visvalingamwyatt as vw
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncEngine, AsyncSession

from virtool.data.topg import compose_legacy_id_multi_expression
from virtool.history.db import patch_otus_to_versions
from virtool.hmm.sql import SQLHMM
from virtool.models.enums import AnalysisWorkflow
from virtool.otus.utils import format_isolate_name

CSV_HEADERS = (
    "OTU",
    "Isolate",
    "Sequence",
    "Length",
    "Weight",
    "Median Depth",
    "Coverage",
)


def calculate_median_depths(hits: list[dict]) -> dict[str, int]:
    """Calculate the median depth for all hits (sequences) in a Pathoscope result
    document.

    :param hits: the pathoscope analysis document to calculate depths for
    :return: a dict of median depths keyed by hit (sequence) ids

    """
    return {hit["id"]: statistics.median(hit["align"]) for hit in hits}


async def format_pathoscope(
    pg: AsyncEngine,
    *,
    results: dict[str, Any],
) -> dict[str, Any]:
    """Format Pathoscope analysis results by retrieving the detected OTUs and
    incorporating them into the returned results.

    Calculate metrics for different organizational levels: OTU, isolate, and sequence.

    Every detected OTU is patched to the version the analysis saw in one batched read,
    on the same collect-then-load shape :func:`format_nuvs` uses for its annotations.
    Reading each OTU as its hits were formatted instead -- a ``patch_to_version`` per
    OTU, fanned out with a ``gather`` -- issued a query and took a pool connection per
    detected OTU, which is enough to saturate both for a result with a few hundred hits.

    :param pg: the application PostgreSQL database object
    :param results: the results to format
    :return: the formatted results

    """
    hits_by_otu = defaultdict(list)

    for hit in results["hits"]:
        otu_id = hit["otu"]["id"]
        otu_version = hit["otu"]["version"]

        hits_by_otu[(otu_id, otu_version)].append(hit)

    patched_otus = await patch_otus_to_versions(pg, hits_by_otu.keys())

    formatted_hits = []

    for otu_specifier, hits in hits_by_otu.items():
        otu_id, _ = otu_specifier
        _, patched_otu = patched_otus[otu_specifier]

        formatted_hits.append(format_pathoscope_hits(otu_id, patched_otu, hits))

    return {**results, "hits": formatted_hits}


def format_pathoscope_hits(
    otu_id: str,
    patched_otu: dict[str, Any],
    hits: list[dict],
):
    max_sequence_length = 0

    for isolate in patched_otu["isolates"]:
        max_sequence_length = max(
            max_sequence_length,
            max(len(s["sequence"]) for s in isolate["sequences"]),
        )

    hits_by_sequence_id = {hit["id"]: hit for hit in hits}

    return {
        "id": otu_id,
        "abbreviation": patched_otu["abbreviation"],
        "name": patched_otu["name"],
        "isolates": list(
            format_pathoscope_isolates(patched_otu["isolates"], hits_by_sequence_id),
        ),
        "length": max_sequence_length,
        "version": patched_otu["version"],
    }


def format_pathoscope_isolates(
    isolates: list[dict[str, Any]],
    hits_by_sequence_ids: dict[str, dict],
) -> list[dict[str, Any]]:
    for isolate in isolates:
        sequences = list(
            format_pathoscope_sequences(isolate["sequences"], hits_by_sequence_ids),
        )

        if any(
            any(key in sequence for sequence in sequences) for key in ("pi", "final")
        ):
            yield {**isolate, "sequences": sequences}


def format_pathoscope_sequences(
    sequences: list[dict[str, Any]],
    hits_by_sequence_id: dict[str, dict],
):
    for sequence in sequences:
        try:
            hit = hits_by_sequence_id[sequence["_id"]]
        except KeyError:
            continue

        try:
            final = hit["final"]
        except KeyError:
            final = {}

        align = hit.get("align")

        if align:
            align = transform_coverage_to_coordinates(align)

        yield {
            "id": sequence["_id"],
            "accession": sequence["accession"],
            "align": align,
            "best": final.get("best", 0),
            "coverage": hit.get("coverage", 0),
            "definition": sequence["definition"],
            "length": len(sequence["sequence"]),
            "pi": final.get("pi", 0),
            "reads": final.get("reads", 0),
        }


async def format_nuvs(
    pg: AsyncEngine,
    *,
    results: dict[str, Any],
) -> dict[str, Any]:
    """Format NuVs analysis results by attaching the HMM annotation data to the hits.

    :param pg: the application PostgreSQL database object
    :param results: the results to format
    :return: the formatted results

    """
    hits = results["hits"]

    hit_ids = list({h["hit"] for s in hits for o in s["orfs"] for h in o["hits"]})

    hmms = {}

    if hit_ids:
        async with AsyncSession(pg) as session:
            rows = (
                await session.execute(
                    select(SQLHMM).where(
                        compose_legacy_id_multi_expression(SQLHMM, hit_ids),
                    ),
                )
            ).scalars()

        for row in rows:
            annotation = {
                "cluster": row.cluster,
                "families": row.families,
                "names": row.names,
            }

            hmms[str(row.id)] = annotation

            if row.legacy_id is not None:
                hmms[row.legacy_id] = annotation

    for sequence in hits:
        for orf in sequence["orfs"]:
            for hit in orf["hits"]:
                hit.update(hmms[str(hit["hit"])])

    return results


async def format_analysis_to_excel(
    pg: AsyncEngine,
    *,
    results: dict[str, Any],
    workflow: str,
    sample_id: str,
) -> bytes:
    """Convert pathoscope analysis results to byte-encoded Excel format for download.

    :param mongo: the database object
    :param pg: the application PostgreSQL database object
    :param results: the results to format
    :param workflow: the analysis workflow
    :param sample_id: the id of the parent sample
    :return: the formatted Excel workbook

    """
    depths = calculate_median_depths(results["hits"])

    formatted = await format_analysis(
        pg,
        workflow=workflow,
        results=results,
    )

    output = io.BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = f"Pathoscope for {sample_id}"

    header_font = openpyxl.styles.Font(name="Calibri", bold=True)

    for index, header in enumerate(CSV_HEADERS):
        col = index + 1
        cell = ws.cell(column=col, row=1, value=header)
        cell.font = header_font

    rows = []

    for otu in formatted["hits"]:
        for isolate in otu["isolates"]:
            for sequence in isolate["sequences"]:
                row = [
                    otu["name"],
                    format_isolate_name(isolate),
                    sequence["accession"],
                    sequence["length"],
                    sequence["pi"],
                    depths.get(sequence["id"], 0),
                    sequence["coverage"],
                ]

                assert len(row) == len(CSV_HEADERS)

                rows.append(row)

    for row_index, row in enumerate(rows):
        row_number = row_index + 2
        for col_index, value in enumerate(row):
            ws.cell(column=col_index + 1, row=row_number, value=value)

    wb.save(output)

    return output.getvalue()


async def format_analysis_to_csv(
    pg: AsyncEngine,
    *,
    results: dict[str, Any],
    workflow: str,
) -> str:
    """Convert pathoscope analysis results to CSV format for download.

    :param pg: the application PostgreSQL database object
    :param results: the results to format
    :param workflow: the analysis workflow
    :return: the formatted CSV data

    """
    depths = calculate_median_depths(results["hits"])

    formatted = await format_analysis(
        pg,
        workflow=workflow,
        results=results,
    )

    output = io.StringIO()

    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)

    writer.writerow(CSV_HEADERS)

    for otu in formatted["hits"]:
        for isolate in otu["isolates"]:
            for sequence in isolate["sequences"]:
                row = [
                    otu["name"],
                    format_isolate_name(isolate),
                    sequence["accession"],
                    sequence["length"],
                    sequence["pi"],
                    depths.get(sequence["id"], 0),
                    sequence["coverage"],
                ]

                writer.writerow(row)

    return output.getvalue()


async def format_analysis(
    pg: AsyncEngine,
    *,
    workflow: str | None,
    results: dict[str, Any],
) -> dict[str, Any]:
    """Format analysis results to be returned by the API.

    :param pg: the application PostgreSQL database object
    :param workflow: the analysis workflow used to dispatch formatting
    :param results: the results to format
    :return: the formatted results

    """
    if workflow is None:
        raise ValueError("Analysis has no workflow field")

    if workflow == AnalysisWorkflow.nuvs.value:
        return await format_nuvs(pg, results=results)

    if "pathoscope" in workflow:
        return await format_pathoscope(pg, results=results)

    raise ValueError(f"Unknown workflow: {workflow}")


def transform_coverage_to_coordinates(
    coverage_list: list[int],
) -> list[tuple[int, int]]:
    """Takes a list of read depths where the list index is equal to the read position
    plus one and returns a list of (x, y) coordinates.

    The coordinates will be simplified using Visvalingham-Wyatt algorithm if the list
    exceeds 100 pairs.

    :param coverage_list: a list of position-indexed depth values
    :return: a list of (x, y) coordinates
    """
    coordinates = [(0, coverage_list[0])]

    last = len(coverage_list) - 1

    for x in range(1, last):
        y = coverage_list[x]
        if y != coverage_list[x - 1] or y != coverage_list[x + 1]:
            coordinates.append((x, y))

    coordinates.append((last, coverage_list[last]))

    if len(coordinates) > 100:
        return vw.simplify(coordinates, ratio=0.4)

    return coordinates
