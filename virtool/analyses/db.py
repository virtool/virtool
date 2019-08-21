import asyncio
import csv
import json
import io
import os
import openpyxl
import openpyxl.styles
import statistics

import aiofiles

import virtool.analyses.utils
import virtool.bio
import virtool.history.db
import virtool.indexes.db
import virtool.jobs.db
import virtool.samples.db
import virtool.db.utils
import virtool.otus.utils
import virtool.utils

CSV_HEADERS = (
    "OTU",
    "Isolate",
    "Sequence",
    "Length",
    "Weight",
    "Median Depth",
    "Coverage"
)

PROJECTION = [
    "_id",
    "algorithm",
    "created_at",
    "ready",
    "job",
    "index",
    "reference",
    "user",
    "sample"
]


def calculate_median_depths(document):
    depths = dict()

    for hit in document["diagnosis"]:
        depths[hit["id"]] = statistics.median(hit["align"])

    return depths


async def format_analysis(db, settings, document):
    """
    Format an analysis document to be returned by the API.

    :param db: the application database client
    :type db: :class:`~motor.motor_asyncio.AsyncIOMotorClient`

    :param settings: the application settings object
    :type settings: :class:`virtool.app_settings.Settings`

    :param document: the analysis document to format
    :type document: dict

    :return: a formatted document
    :rtype: dict

    """
    algorithm = document.get("algorithm", None)

    if algorithm == "nuvs":
        return await format_nuvs(db, settings, document)

    if algorithm and "pathoscope" in algorithm:
        return await format_pathoscope(db, settings, document)

    raise ValueError("Could not determine analysis algorithm")


async def format_analysis_to_csv(db, settings, document):
    depths = calculate_median_depths(document)

    formatted = await format_analysis(db, settings, document)

    output = io.StringIO()

    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)

    writer.writerow(CSV_HEADERS)

    for otu in formatted["diagnosis"]:
        for isolate in otu["isolates"]:
            for sequence in isolate["sequences"]:
                row = [
                    otu["name"],
                    virtool.otus.utils.format_isolate_name(isolate),
                    sequence["accession"],
                    sequence["length"],
                    sequence["pi"],
                    depths.get(sequence["id"], 0),
                    sequence["coverage"]
                ]

                writer.writerow(row)

    return output.getvalue()


async def format_analysis_to_excel(db, settings, document):
    depths = calculate_median_depths(document)

    formatted = await format_analysis(db, settings, document)

    output = io.BytesIO()

    wb = openpyxl.Workbook()
    ws = wb.active

    ws.title = f"Pathoscope for {document['sample']['id']}"

    header_font = openpyxl.styles.Font(name="Calibri", bold=True)

    for index, header in enumerate(CSV_HEADERS):
        col = index + 1
        cell = ws.cell(column=col, row=1, value=header)
        cell.font = header_font

    rows = list()

    for otu in formatted["diagnosis"]:
        for isolate in otu["isolates"]:
            for sequence in isolate["sequences"]:
                row = [
                    otu["name"],
                    virtool.otus.utils.format_isolate_name(isolate),
                    sequence["accession"],
                    sequence["length"],
                    sequence["pi"],
                    depths.get(sequence["id"], 0),
                    sequence["coverage"]
                ]

                assert len(row) == len(CSV_HEADERS)

                rows.append(row)

    for row_index, row in enumerate(rows):
        row_number = row_index + 2
        for col_index, value in enumerate(row):
            ws.cell(column=col_index + 1, row=row_number, value=value)

    wb.save(output)

    return output.getvalue()


async def format_nuvs(db, settings, document):
    if document["results"] == "file":
        path = virtool.analyses.utils.join_nuvs_json_path(
            settings["data_path"],
            document["_id"],
            document["sample"]["id"]
        )

        async with aiofiles.open(path, "r") as f:
            json_string = await f.read()
            document["results"] = dict(document, results=json.loads(json_string))

    hit_ids = list({h["hit"] for s in document["results"] for o in s["orfs"] for h in o["hits"]})

    cursor = db.hmm.find({"_id": {"$in": hit_ids}}, ["cluster", "families", "names"])

    hmms = {d.pop("_id"): d async for d in cursor}

    for sequence in document["results"]:
        for orf in sequence["orfs"]:
            for hit in orf["hits"]:
                hit.update(hmms[hit["hit"]])

    return document


async def format_pathoscope(db, settings, document):
    if document["diagnosis"] == "file":
        path = virtool.analyses.utils.join_pathoscope_json_path(
            settings["data_path"],
            document["_id"],
            document["sample"]["id"]
        )

        async with aiofiles.open(path, "r") as f:
            json_string = await f.read()
            document.update(json.loads(json_string))

    formatted = dict()

    otu_specifiers = {(hit["otu"]["id"], hit["otu"]["version"]) for hit in document["diagnosis"]}

    patched_otus = await asyncio.gather(*[
        virtool.history.db.patch_to_version(db, otu_id, version) for otu_id, version in otu_specifiers
    ])

    patched_otus = {patched["_id"]: patched for _, patched, _ in patched_otus}

    for hit in document["diagnosis"]:

        otu_id = hit["otu"]["id"]

        otu_document = patched_otus[otu_id]

        max_ref_length = 0

        for isolate in otu_document["isolates"]:
            max_ref_length = max(max_ref_length, max([len(s["sequence"]) for s in isolate["sequences"]]))

        otu = {
            "id": otu_id,
            "name": otu_document["name"],
            "version": otu_document["version"],
            "abbreviation": otu_document["abbreviation"],
            "isolates": otu_document["isolates"],
            "length": max_ref_length
        }

        formatted[otu_id] = otu

        for isolate in otu["isolates"]:
            for sequence in isolate["sequences"]:
                if sequence["_id"] == hit["id"]:
                    sequence.update(hit)
                    sequence["length"] = len(sequence["sequence"])

                    del sequence["otu"]
                    del sequence["otu_id"]
                    del sequence["isolate_id"]

    document["diagnosis"] = [formatted[otu_id] for otu_id in formatted]

    for otu in document["diagnosis"]:
        for isolate in list(otu["isolates"]):
            if not any((key in sequence for sequence in isolate["sequences"]) for key in ("pi", "final")):
                otu["isolates"].remove(isolate)
                continue

            for sequence in isolate["sequences"]:
                if "final" in sequence:
                    sequence.update(sequence.pop("final"))
                    del sequence["initial"]
                if "pi" not in sequence:
                    sequence.update({
                        "pi": 0,
                        "reads": 0,
                        "coverage": 0,
                        "best": 0,
                        "length": len(sequence["sequence"])
                    })

                sequence["id"] = sequence.pop("_id")

                if "align" in sequence:
                    sequence["align"] = virtool.analyses.utils.transform_coverage_to_coordinates(sequence["align"])

                del sequence["sequence"]

    return document


async def new(app, sample_id, ref_id, user_id, algorithm):
    """
    Creates a new analysis. Ensures that a valid subtraction host was the submitted. Configures read and write
    permissions on the sample document and assigns it a creator username based on the requesting connection.

    """
    db = app["db"]
    settings = app["settings"]

    # Get the current id and version of the otu index currently being used for analysis.
    index_id, index_version = await virtool.indexes.db.get_current_id_and_version(db, ref_id)

    sample = await db.samples.find_one(sample_id, ["name"])

    analysis_id = await virtool.db.utils.get_new_id(db.analyses)

    job_id = await virtool.db.utils.get_new_id(db.jobs)

    document = {
        "_id": analysis_id,
        "ready": False,
        "created_at": virtool.utils.timestamp(),
        "job": {
            "id": job_id
        },
        "algorithm": algorithm,
        "sample": {
            "id": sample_id
        },
        "index": {
            "id": index_id,
            "version": index_version
        },
        "reference": {
            "id": ref_id,
            "name": await virtool.db.utils.get_one_field(db.references, "name", ref_id)
        },
        "user": {
            "id": user_id,
        }
    }

    task_args = {
        "analysis_id": analysis_id,
        "ref_id": ref_id,
        "sample_id": sample_id,
        "sample_name": sample["name"],
        "index_id": index_id
    }

    await db.analyses.insert_one(document)

    # Create job document.
    job = await virtool.jobs.db.create(
        db,
        settings,
        document["algorithm"],
        task_args,
        user_id
    )

    await app["jobs"].enqueue(job["_id"])

    await virtool.samples.db.recalculate_algorithm_tags(db, sample_id)

    return document


async def update_nuvs_blast(db, settings, analysis_id, sequence_index, rid):
    """
    Update the BLAST data for a sequence in a NuVs analysis.

    :param settings:
    :param db:

    :param analysis_id:
    :param sequence_index:
    :param rid:

    :return: the blast data and the complete analysis document
    :rtype: Tuple[dict, dict]

    """
    # Do initial check of RID to populate BLAST embedded document.
    data = {
        "rid": rid,
        "ready": await virtool.bio.check_rid(settings, rid),
        "last_checked_at": virtool.utils.timestamp(),
        "interval": 3
    }

    document = await db.analyses.find_one_and_update({"_id": analysis_id, "results.index": sequence_index}, {
        "$set": {
            "results.$.blast": data
        }
    })

    return data, document


async def remove_orphaned_directories(app):
    """
    Remove all analysis directories for which an analysis document does not exist in the database.

    :param app:
    """
    db = app["db"]

    samples_path = os.path.join(app["settings"]["data_path"], "samples")

    existing_ids = set(await db.analyses.distinct("_id"))

    for sample_id in os.listdir(samples_path):
        analyses_path = os.path.join(samples_path, sample_id, "analysis")

        to_delete = set(os.listdir(analyses_path)) - existing_ids

        for analysis_id in to_delete:
            analysis_path = os.path.join(analyses_path, analysis_id)
            try:
                await app["run_in_thread"](virtool.utils.rm, analysis_path, True)
            except FileNotFoundError:
                pass
